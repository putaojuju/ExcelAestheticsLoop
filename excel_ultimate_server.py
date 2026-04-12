"""
通用 Excel 安全读写与视觉审计 MCP 服务器 (V2.1-Vision)
======================================================
为 E:/lin 工作区提供 Excel (.xlsx) 文件的安全读写及「像素级视觉审计」能力。
核心场景：基于下单图制作石材开料清单。

安全机制（源自 OWASP 公式注入防御 + Deep Research 成果）：
  - 写前备份（.bak）
  - OWASP 危险前缀清洗
  - 写后验证读回
  - 所有写操作返回变更摘要供人工审查

视觉审计能力 (V2.0+ 新增)：
  - smart_audit_read: 读取数据的同时自动捕获高危列视觉快照
  - get_visual_slice: 对指定区域进行高清 PNG 渲染 (指哪打哪)
  - get_layout_metrics: 扫描合并单元格和列宽度（解决空间盲区）

依赖：pip install fastmcp openpyxl pywin32 Pillow
"""

from fastmcp import FastMCP
import openpyxl
from openpyxl.utils import get_column_letter
import os
import shutil
import json
import io
import sys
import time
from datetime import datetime
import win32com.client
from PIL import ImageGrab

# 强制 UTF-8 输出
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

mcp = FastMCP("Excel Ultimate Vision Plane")

# 默认模板路径
DEFAULT_TEMPLATE = os.environ.get("EXCEL_TEMPLATE", r"C:\path\to\template.xlsx")

# 视觉审计配置
CACHE_DIR = os.environ.get("EXCEL_CACHE_DIR", os.path.join(os.path.expanduser("~"), ".excel_mcp_cache"))
AUDIT_KEYWORDS = ["备注", "remarks", "附件", "attachments", "项目", "description", "对应单号", "audit", "审计"]

# ============================================================
# OWASP 公式注入防御层
# ============================================================
DANGEROUS_PREFIXES = ('=', '+', '-', '@', '\t', '\r', '\n',
                      '＝', '＋', '－', '＠')

def sanitize_cell_value(value):
    """
    OWASP 标准：阻断公式注入。
    数值类型直接通过；字符串类型检查危险前缀，
    发现则加前置单引号强制文本模式。
    """
    if value is None:
        return value
    if isinstance(value, (int, float)):
        return value
    s = str(value)
    if s and s[0] in DANGEROUS_PREFIXES:
        return "'" + s
    return value


def _make_backup(filepath: str) -> str:
    """写操作前自动备份，返回备份路径。"""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(filepath)
    bak_path = f"{base}.{ts}.bak{ext}"
    shutil.copy2(filepath, bak_path)
    return bak_path


def _col_letter_to_index(letter: str) -> int:
    """将列字母 (A, B, ..., U) 转为 1-indexed 数字。"""
    letter = letter.upper().strip()
    result = 0
    for ch in letter:
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result


def _resolve_col(col) -> int:
    """接受列字母或数字，统一返回 1-indexed 列号。"""
    if isinstance(col, int):
        return col
    if isinstance(col, str):
        if col.isdigit():
            return int(col)
        return _col_letter_to_index(col)
    return int(col)


def _get_col_letter_from_int(n: int) -> str:
    """Int to Column Letter (1 -> A)."""
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


def _get_primary_cell(ws, row, col_letter):
    """Bypasses MergedCell read-only constraints by routing to the top-left cell."""
    coord = f"{col_letter}{row}"
    for merged_range in ws.merged_cells.ranges:
        if coord in merged_range:
            return merged_range.start_cell.coordinate
    return coord


# ============================================================
# 核心视觉引擎 (Native RIP) — V2.0 新增
# ============================================================
def _internal_render(excel_path, sheet_name, range_str, output_path):
    """
    核心渲染逻辑：使用 xlScreen + xlBitmap 绕过剪贴板权限与 headless 限制。
    """
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = True
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(os.path.abspath(excel_path), ReadOnly=True)
        ws = wb.Sheets(sheet_name)
        ws.Activate()

        target_range = ws.Range(range_str)
        target_range.Select()

        # 尝试 3 次 CopyPicture (应对剪贴板偶发忙)
        success = False
        for _ in range(3):
            try:
                target_range.CopyPicture(Appearance=1, Format=2)
                time.sleep(0.5)
                img = ImageGrab.grabclipboard()
                if img:
                    img.save(output_path, "PNG")
                    success = True
                    break
            except:
                time.sleep(1)
        return success
    finally:
        try:
            wb.Close(SaveChanges=False)
            excel.Quit()
        except:
            pass


# ============================================================
# 工具 1：读取 Excel 结构
# ============================================================
@mcp.tool()
def read_excel_schema(file_path: str) -> str:
    """
    读取指定 Excel 文件的结构概览：所有 Sheet 名、各 Sheet 行数/列数/表头。
    用于在写入前了解文件结构。
    """
    if not os.path.exists(file_path):
        return f"Error: 文件不存在 → {file_path}"

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        result = {"file": os.path.basename(file_path), "sheets": []}

        for sname in wb.sheetnames:
            ws = wb[sname]
            # 读取前两行作为表头示例
            headers = []
            for r in range(1, min(ws.max_row + 1, 6)):
                row_data = {}
                for c in range(1, min(ws.max_column + 1, 22)):
                    v = ws.cell(row=r, column=c).value
                    if v is not None:
                        row_data[f"Col{c}"] = str(v)[:40]
                if row_data:
                    headers.append({"row": r, "cells": row_data})

            result["sheets"].append({
                "name": sname,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "header_sample": headers
            })

        wb.close()
        return json.dumps(result, ensure_ascii=False, indent=2)
    except Exception as e:
        return f"Error: 读取失败 → {str(e)}"


# ============================================================
# 工具 2：读取指定行范围
# ============================================================
@mcp.tool()
def read_excel_rows(file_path: str, sheet_name: str,
                    start_row: int = 1, end_row: int = 50) -> str:
    """
    读取指定 Sheet 的指定行范围，返回 JSON 格式数据。
    适合查看已有清单内容或验证写入结果。
    """
    if not os.path.exists(file_path):
        return f"Error: 文件不存在 → {file_path}"

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' 不存在。可选: {wb.sheetnames}"

        ws = wb[sheet_name]
        rows = []
        for r in range(start_row, min(end_row + 1, ws.max_row + 1)):
            row_data = {}
            for c in range(1, min(ws.max_column + 1, 22)):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    row_data[f"Col{c}"] = v if isinstance(v, (int, float)) else str(v)[:60]
            if row_data:
                rows.append({"row": r, "cells": row_data})

        wb.close()
        return json.dumps(rows, ensure_ascii=False, indent=2)
    except Exception as e:
        return f"Error: 读取失败 → {str(e)}"


# ============================================================
# 工具 3：从模板复制新清单
# ============================================================
@mcp.tool()
def copy_template(dest_path: str, template_path: str = "") -> str:
    """
    从模板文件复制生成新的清单文件。
    如不指定 template_path，使用默认模板。
    dest_path 为新文件的完整路径。
    """
    src = template_path if template_path else DEFAULT_TEMPLATE

    if not os.path.exists(src):
        return f"Error: 模板不存在 → {src}"
    if os.path.exists(dest_path):
        return f"Error: 目标文件已存在 → {dest_path}（避免意外覆盖）"

    try:
        # 确保目标目录存在
        dest_dir = os.path.dirname(dest_path)
        if dest_dir:
            os.makedirs(dest_dir, exist_ok=True)

        shutil.copy2(src, dest_path)
        return f"OK: 已从模板创建新文件\n  模板: {src}\n  新文件: {dest_path}"
    except Exception as e:
        return f"Error: 复制失败 → {str(e)}"


# ============================================================
# 工具 4：预览写入（不实际修改文件）
# ============================================================
@mcp.tool()
def preview_write(file_path: str, sheet_name: str,
                  updates: str) -> str:
    """
    预览写入效果，不实际修改文件。返回每个单元格的 旧值→新值 对比。

    updates 参数为 JSON 字符串，格式：
    [
      {"row": 6, "col": "B", "value": "A1,A28~A31"},
      {"row": 6, "col": "C", "value": "山西黑"},
      {"row": 6, "col": "E", "value": 603},
      {"row": 6, "col": "F", "value": 320}
    ]
    col 可以是字母 (A-U) 或数字 (1-21)。
    """
    if not os.path.exists(file_path):
        return f"Error: 文件不存在 → {file_path}"

    try:
        update_list = json.loads(updates)
    except json.JSONDecodeError as e:
        return f"Error: updates JSON 解析失败 → {str(e)}"

    if len(update_list) > 200:
        return f"Error: 单次更新不得超过200个单元格（当前: {len(update_list)}）"

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' 不存在。可选: {wb.sheetnames}"

        ws = wb[sheet_name]
        preview = []

        for u in update_list:
            row = int(u["row"])
            col = _resolve_col(u["col"])
            new_val = u["value"]
            old_val = ws.cell(row=row, column=col).value

            # 对新值执行 OWASP 清洗（预览中也标注）
            sanitized = sanitize_cell_value(new_val)
            was_sanitized = sanitized != new_val

            preview.append({
                "cell": f"R{row}C{col}",
                "old": str(old_val) if old_val is not None else "(空)",
                "new": str(sanitized),
                "sanitized": was_sanitized
            })

        wb.close()

        summary = f"预览: {len(preview)} 个单元格将被修改\n"
        summary += f"文件: {os.path.basename(file_path)} → Sheet '{sheet_name}'\n"
        summary += "-" * 50 + "\n"
        for p in preview:
            flag = " ⚠️已清洗" if p["sanitized"] else ""
            summary += f"  {p['cell']}: {p['old']} → {p['new']}{flag}\n"

        return summary
    except Exception as e:
        return f"Error: 预览失败 → {str(e)}"


# ============================================================
# 工具 5：确认写入（带备份+清洗+验证+日志）
# ============================================================
@mcp.tool()
def commit_write(file_path: str, sheet_name: str,
                 updates: str, reason: str = "") -> str:
    """
    确认写入：执行完整的四步防御链。
    Step 1: 备份原文件 (.bak)
    Step 2: OWASP 清洗所有待写入值
    Step 3: openpyxl 写入 + 即时验证读回
    Step 4: 输出变更摘要日志

    updates 格式同 preview_write。
    reason: 写入原因（记录到日志）。
    """
    if not os.path.exists(file_path):
        return f"Error: 文件不存在 → {file_path}"

    try:
        update_list = json.loads(updates)
    except json.JSONDecodeError as e:
        return f"Error: updates JSON 解析失败 → {str(e)}"

    if len(update_list) > 200:
        return f"Error: 单次更新不得超过200个单元格（当前: {len(update_list)}）"

    # Step 1: 备份
    bak_path = _make_backup(file_path)

    try:
        # Step 2 & 3: 清洗 + 写入（含 MergedCell 路由）
        wb = openpyxl.load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' 不存在。可选: {wb.sheetnames}"

        ws = wb[sheet_name]
        log_entries = []

        for u in update_list:
            row = int(u["row"])
            col = _resolve_col(u["col"])
            new_val = u["value"]
            col_letter = get_column_letter(col)

            # MergedCell 智能路由
            target_coord = _get_primary_cell(ws, row, col_letter)
            old_val = ws[target_coord].value
            sanitized_val = sanitize_cell_value(new_val)

            # 写入
            ws[target_coord] = sanitized_val

            log_entries.append({
                "cell": f"{col_letter}{row}→{target_coord}",
                "old": str(old_val) if old_val is not None else "(空)",
                "new": str(sanitized_val),
                "sanitized": sanitized_val != new_val
            })

        wb.save(file_path)
        wb.close()

        # Step 3 (续): 验证读回
        wb_verify = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws_verify = wb_verify[sheet_name]
        verify_errors = []

        for u in update_list:
            row = int(u["row"])
            col = _resolve_col(u["col"])
            expected = sanitize_cell_value(u["value"])
            actual = ws_verify.cell(row=row, column=col).value

            # 宽松比较（数值精度和类型差异）
            if str(expected) != str(actual) and expected != actual:
                verify_errors.append(f"  R{row}C{col}: 期望={expected}, 实际={actual}")

        wb_verify.close()

        # Step 4: 生成变更摘要
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        summary = f"✅ 写入成功 — {ts}\n"
        summary += f"文件: {os.path.basename(file_path)} → Sheet '{sheet_name}'\n"
        summary += f"备份: {os.path.basename(bak_path)}\n"
        if reason:
            summary += f"原因: {reason}\n"
        summary += f"修改: {len(log_entries)} 个单元格\n"
        summary += "-" * 50 + "\n"

        for entry in log_entries:
            flag = " ⚠️已清洗" if entry["sanitized"] else ""
            summary += f"  {entry['cell']}: {entry['old']} → {entry['new']}{flag}\n"

        if verify_errors:
            summary += "\n❌ 验证异常:\n" + "\n".join(verify_errors)
        else:
            summary += "\n✅ 写后验证通过"

        return summary

    except Exception as e:
        return f"Error: 写入失败 → {str(e)}\n备份文件: {bak_path}（可用于恢复）"


# ============================================================
# 工具 6：批量追加行
# ============================================================
@mcp.tool()
def append_rows(file_path: str, sheet_name: str,
                start_row: int, rows_data: str,
                reason: str = "") -> str:
    """
    从指定行开始，批量写入多行数据（典型场景：填入下料单明细）。

    rows_data 为 JSON 字符串，格式：
    [
      {"B": "A1", "C": "山西黑", "D": "烧面", "E": 603, "F": 320, "G": 30, "H": 1},
      {"B": "A2", "C": "山西黑", "D": "烧面", "E": 615, "F": 305, "G": 30, "H": 8}
    ]
    列名用字母标识。自动为每行计算序号 (A列) 和面积 (I列 = E*F/1000000*H)。
    """
    if not os.path.exists(file_path):
        return f"Error: 文件不存在 → {file_path}"

    try:
        data_rows = json.loads(rows_data)
    except json.JSONDecodeError as e:
        return f"Error: rows_data JSON 解析失败 → {str(e)}"

    if len(data_rows) > 200:
        return f"Error: 单次不得超过200行（当前: {len(data_rows)}）"

    # 构造 updates 列表，转换为 commit_write 所需格式
    updates = []
    for i, row_data in enumerate(data_rows):
        current_row = start_row + i

        # 自动填序号 (A列 = col 1)
        updates.append({"row": current_row, "col": 1, "value": i + 1})

        # 写入各列数据
        for col_letter, value in row_data.items():
            col_idx = _resolve_col(col_letter)
            updates.append({"row": current_row, "col": col_idx, "value": value})

        # 自动计算面积 (I列 = E*F/1000000*H)
        e_val = row_data.get("E", row_data.get("e", 0))
        f_val = row_data.get("F", row_data.get("f", 0))
        h_val = row_data.get("H", row_data.get("h", 1))
        try:
            area = float(e_val) * float(f_val) / 1000000.0 * float(h_val)
            area = round(area, 6)
            updates.append({"row": current_row, "col": 9, "value": area})  # I列 = col 9
        except (ValueError, TypeError):
            pass  # 如果尺寸不是数值，跳过面积计算

    # 委托给 commit_write 执行（含完整防御链）
    updates_json = json.dumps(updates, ensure_ascii=False)
    return commit_write(file_path, sheet_name, updates_json,
                       reason=reason or f"批量追加 {len(data_rows)} 行下料数据")


# ============================================================
# 工具 7：读取并提取 Excel 中的图片（带坐标锚定）
# ============================================================
@mcp.tool()
def read_excel_images(file_path: str, sheet_name: str, output_dir: str) -> str:
    """
    提取指定 Sheet 中的所有悬浮图片，并返回其归属的单元格坐标。
    图片将以 'img_R{行}C{列}_{序号}.png' 的格式保存在 output_dir。
    """
    if not os.path.exists(file_path):
        return f"Error: 文件不存在 → {file_path}"

    if not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    try:
        # 必须使用 openpyxl 加载，注意：read_only 模式下可能无法获取 _images
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' 不存在。"

        ws = wb[sheet_name]
        extracted_images = []

        if not hasattr(ws, '_images') or not ws._images:
            wb.close()
            return json.dumps({"count": 0, "images": [], "msg": "未发现图片"}, ensure_ascii=False)

        for i, img in enumerate(ws._images):
            # 获取锚点位置 (0-indexed -> 1-indexed)
            row = img.anchor._from.row + 1
            col = img.anchor._from.col + 1

            # 提取图片后缀
            ext = ".png"
            if hasattr(img, 'format') and img.format:
                ext = f".{img.format.lower()}"

            filename = f"img_R{row}C{col}_{i}{ext}"
            target_path = os.path.join(output_dir, filename)

            # 写入二进制数据
            with open(target_path, "wb") as f:
                f.write(img._data())

            extracted_images.append({
                "index": i,
                "row": row,
                "col": col,
                "local_path": target_path,
                "filename": filename
            })

        wb.close()
        return json.dumps({
            "count": len(extracted_images),
            "images": extracted_images,
            "msg": f"成功提取 {len(extracted_images)} 张图片"
        }, ensure_ascii=False, indent=2)

    except Exception as e:
        return f"Error: 图片提取失败 → {str(e)}"


# ============================================================
# 工具 8：智能审计读取 — 文图合一 (V2.0 新增)
# ============================================================
@mcp.tool()
def smart_audit_read(file_path: str, sheet_name: str) -> str:
    """
    读取指定 Sheet 的 Schema，并自动捕获「高危列 (备注/单号/附件)」的高清视觉快照。
    返回 JSON 包含数据结构及生成的图片路径。
    """
    if not os.path.exists(file_path):
        return f"Error: 文件不存在 → {file_path}"

    # 确保缓存目录存在
    if not os.path.exists(CACHE_DIR):
        os.makedirs(CACHE_DIR, exist_ok=True)

    ts = datetime.now().strftime("%H%M%S")
    report = {
        "file": os.path.basename(file_path),
        "sheet": sheet_name,
        "timestamp": ts,
        "columns": [],
        "visual_cache": []
    }

    try:
        # 1. 文本层面扫描 (使用 openpyxl 保证速度)
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' 不存在。"

        ws = wb[sheet_name]
        # 自动探测表头 (通常在前 5 行)
        header_row = 1
        for r in range(1, 6):
            if any(ws.cell(row=r, column=c).value for c in range(1, 5)):
                header_row = r
                break

        for c in range(1, ws.max_column + 1):
            h_val = ws.cell(row=header_row, column=c).value
            h_str = str(h_val) if h_val else f"Column_{c}"
            col_letter = _get_col_letter_from_int(c)

            needs_v = any(k in h_str.lower() for k in AUDIT_KEYWORDS)
            report["columns"].append({
                "letter": col_letter,
                "header": h_str,
                "needs_visual": needs_v
            })
        wb.close()

        # 2. 视觉层面 RIP (针对高危列)
        for col in report["columns"]:
            if col["needs_visual"]:
                letter = col["letter"]
                safe_h = "".join([c for c in col['header'] if c.isalnum()])[:10]
                out_name = f"audit_{ts}_{letter}_{safe_h}.png"
                out_path = os.path.join(CACHE_DIR, out_name)

                # 渲染前 30 行
                if _internal_render(file_path, sheet_name, f"{letter}1:{letter}30", out_path):
                    col["visual_cache_path"] = out_path
                    report["visual_cache"].append(out_path)

        return json.dumps(report, ensure_ascii=False, indent=2)
    except Exception as e:
        return f"Error: 审计读取失败 → {str(e)}"


# ============================================================
# 工具 9：指哪打哪视觉切片 (V2.0 新增)
# ============================================================
@mcp.tool()
def get_visual_slice(file_path: str, sheet_name: str, range_str: str) -> str:
    """
    对 Excel 指定区域进行高清截图渲染。
    range_str 可以是 'B1:C20' 或单列字母 'O'。
    """
    if not os.path.exists(CACHE_DIR):
        os.makedirs(CACHE_DIR, exist_ok=True)

    # 智能补全
    if ":" not in range_str and len(range_str) <= 3:
        range_str = f"{range_str}1:{range_str}40"

    ts = datetime.now().strftime("%H%M%S")
    out_path = os.path.join(CACHE_DIR, f"rip_{ts}_{range_str.replace(':', '_')}.png")

    if _internal_render(file_path, sheet_name, range_str, out_path):
        return f"OK: 视觉切片已生成 → {out_path}"
    else:
        return "Error: 渲染失败，请检查 Excel 状态或区域有效性。"


# ============================================================
# 工具 10：布局度量检查器 (来自 Ultimate 版)
# ============================================================
@mcp.tool()
def get_layout_metrics(file_path: str, sheet_name: str) -> str:
    """[The Inspector] 扫描合并单元格和列宽度，用于解决排版中的「空间盲区」问题。"""
    if not os.path.exists(file_path):
        return f"Error: 文件不存在 → {file_path}"
    try:
        # 修正：merged_cells 不支持 read_only 模式
        wb = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{sheet_name}' 不存在。"
        ws = wb[sheet_name]
        merged = [str(r) for r in ws.merged_cells.ranges]
        metrics = {
            "merged_cells": merged,
            "column_widths": {get_column_letter(i): ws.column_dimensions[get_column_letter(i)].width
                              for i in range(1, ws.max_column + 1)}
        }
        wb.close()
        return json.dumps(metrics, ensure_ascii=False, indent=2)
    except Exception as e:
        return f"Error: {str(e)}"


# ============================================================
if __name__ == "__main__":
    mcp.run()
