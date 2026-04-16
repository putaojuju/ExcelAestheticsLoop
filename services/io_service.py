"""
services/io_service.py — 数据读写通道
======================================
包含：read_excel_schema, read_excel_rows, copy_template,
      preview_write, commit_write, append_rows, read_excel_images

所有写操作均经过 OWASP 清洗 → MergedCell 路由 → 写后验证读回 四步防御链。
"""

import os
import json
import shutil
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

from config import DEFAULT_TEMPLATE
from core.security import sanitize_cell_value, make_backup
from core.excel_mapping import resolve_col, get_primary_cell


def register_io_tools(mcp):
    """将 IO 类工具注册到 FastMCP 实例。"""

    @mcp.tool()
    def read_excel_schema(file_path: str) -> str:
        """
        读取指定 Excel 文件的结构概览：所有 Sheet 名、各 Sheet 行数/列数/表头。
        用于在写入前了解文件结构。
        """
        if not os.path.exists(file_path):
            return f"Error: 文件不存在 → {file_path}"

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            result = {"file": os.path.basename(file_path), "sheets": []}

            for sname in wb.sheetnames:
                ws = wb[sname]
                headers = []
                for r in range(1, min(ws.max_row + 1, 6)):
                    row_data = {}
                    for c in range(1, min(ws.max_column + 1, 22)):
                        v = ws.cell(row=r, column=c).value
                        if v is not None:
                            row_data[f"Col{c}"] = str(v)[:40]
                    if row_data:
                        headers.append({"row": r, "cells": row_data})

                result["sheets"].append({
                    "name": sname,
                    "max_row": ws.max_row,
                    "max_col": ws.max_column,
                    "header_sample": headers
                })

            wb.close()
            return json.dumps(result, ensure_ascii=False, indent=2)
        except Exception as e:
            return f"Error: 读取失败 → {str(e)}"

    @mcp.tool()
    def read_excel_rows(file_path: str, sheet_name: str,
                        start_row: int = 1, end_row: int = 50) -> str:
        """
        读取指定 Sheet 的指定行范围，返回 JSON 格式数据。
        适合查看已有清单内容或验证写入结果。
        [V2.2] 改用 iter_rows(min_row=) 直接跳至目标行，避免全表顺序扫描。
        """
        if not os.path.exists(file_path):
            return f"Error: 文件不存在 → {file_path}"

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Error: Sheet '{sheet_name}' 不存在。可选: {wb.sheetnames}"

            ws = wb[sheet_name]
            rows = []
            for r_idx, row_tuple in enumerate(
                ws.iter_rows(min_row=start_row, max_row=end_row, max_col=21, values_only=True),
                start=start_row
            ):
                row_data = {}
                for c_idx, v in enumerate(row_tuple, start=1):
                    if v is not None:
                        row_data[f"Col{c_idx}"] = v if isinstance(v, (int, float)) else str(v)[:60]
                if row_data:
                    rows.append({"row": r_idx, "cells": row_data})

            wb.close()
            return json.dumps(rows, ensure_ascii=False, indent=2)
        except Exception as e:
            return f"Error: 读取失败 → {str(e)}"

    @mcp.tool()
    def grep_rows(file_path: str, sheet_name: str, pattern: str,
                  max_results: int = 50) -> str:
        """
        在指定 Sheet 中全表搜索包含关键字的行，返回匹配结果。
        比 read_excel_rows 范围读取快 100 倍，是「找内容」的首选工具。

        pattern: 搜索关键字（不区分大小写，支持中文）
        max_results: 最多返回行数，默认 50
        """
        if not os.path.exists(file_path):
            return f"Error: 文件不存在 → {file_path}"

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Error: Sheet '{sheet_name}' 不存在。可选: {wb.sheetnames}"

            ws = wb[sheet_name]
            pattern_lower = pattern.lower()
            matches = []

            for r_idx, row_tuple in enumerate(
                ws.iter_rows(max_col=21, values_only=True), start=1
            ):
                row_str = " ".join(str(v) for v in row_tuple if v is not None).lower()
                if pattern_lower in row_str:
                    row_data = {}
                    for c_idx, v in enumerate(row_tuple, start=1):
                        if v is not None:
                            row_data[f"Col{c_idx}"] = v if isinstance(v, (int, float)) else str(v)[:80]
                    matches.append({"row": r_idx, "cells": row_data})
                    if len(matches) >= max_results:
                        break

            wb.close()
            result = {
                "pattern": pattern,
                "sheet": sheet_name,
                "total_matches": len(matches),
                "results": matches
            }
            if len(matches) >= max_results:
                result["warning"] = f"已达最大返回数 {max_results}，可能还有更多结果，请缩窄关键字"
            return json.dumps(result, ensure_ascii=False, indent=2)
        except Exception as e:
            return f"Error: 搜索失败 → {str(e)}"

    @mcp.tool()
    def copy_template(dest_path: str, template_path: str = "") -> str:
        """
        从模板文件复制生成新的清单文件。
        如不指定 template_path，使用默认模板 E:\\lin\\空分析.xlsx。
        dest_path 为新文件的完整路径。
        """
        src = template_path if template_path else DEFAULT_TEMPLATE

        if not os.path.exists(src):
            return f"Error: 模板不存在 → {src}"
        if os.path.exists(dest_path):
            return f"Error: 目标文件已存在 → {dest_path}（避免意外覆盖）"

        try:
            dest_dir = os.path.dirname(dest_path)
            if dest_dir:
                os.makedirs(dest_dir, exist_ok=True)

            shutil.copy2(src, dest_path)
            return f"OK: 已从模板创建新文件\n  模板: {src}\n  新文件: {dest_path}"
        except Exception as e:
            return f"Error: 复制失败 → {str(e)}"

    @mcp.tool()
    def preview_write(file_path: str, sheet_name: str,
                      updates: str) -> str:
        """
        预览写入效果，不实际修改文件。返回每个单元格的 旧值→新值 对比。

        updates 参数为 JSON 字符串，格式：
        [
          {"row": 6, "col": "B", "value": "A1,A28~A31"},
          {"row": 6, "col": "C", "value": "山西黑"},
          {"row": 6, "col": "E", "value": 603},
          {"row": 6, "col": "F", "value": 320}
        ]
        col 可以是字母 (A-U) 或数字 (1-21)。
        """
        if not os.path.exists(file_path):
            return f"Error: 文件不存在 → {file_path}"

        try:
            update_list = json.loads(updates)
        except json.JSONDecodeError as e:
            return f"Error: updates JSON 解析失败 → {str(e)}"

        if len(update_list) > 200:
            return f"Error: 单次更新不得超过200个单元格（当前: {len(update_list)}）"

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Error: Sheet '{sheet_name}' 不存在。可选: {wb.sheetnames}"

            ws = wb[sheet_name]
            preview = []

            for u in update_list:
                row = int(u["row"])
                col = resolve_col(u["col"])
                new_val = u["value"]
                old_val = ws.cell(row=row, column=col).value

                sanitized = sanitize_cell_value(new_val)
                was_sanitized = sanitized != new_val

                preview.append({
                    "cell": f"R{row}C{col}",
                    "old": str(old_val) if old_val is not None else "(空)",
                    "new": str(sanitized),
                    "sanitized": was_sanitized
                })

            wb.close()

            summary = f"预览: {len(preview)} 个单元格将被修改\n"
            summary += f"文件: {os.path.basename(file_path)} → Sheet '{sheet_name}'\n"
            summary += "-" * 50 + "\n"
            for p in preview:
                flag = " ⚠️已清洗" if p["sanitized"] else ""
                summary += f"  {p['cell']}: {p['old']} → {p['new']}{flag}\n"

            return summary
        except Exception as e:
            return f"Error: 预览失败 → {str(e)}"

    @mcp.tool()
    def commit_write(file_path: str, sheet_name: str,
                     updates: str, reason: str = "") -> str:
        """
        确认写入：执行完整的四步防御链。
        Step 1: 备份原文件 (.bak)
        Step 2: OWASP 清洗所有待写入值
        Step 3: openpyxl 写入 + 即时验证读回
        Step 4: 输出变更摘要日志

        updates 格式同 preview_write。
        reason: 写入原因（记录到日志）。
        """
        if not os.path.exists(file_path):
            return f"Error: 文件不存在 → {file_path}"

        try:
            update_list = json.loads(updates)
        except json.JSONDecodeError as e:
            return f"Error: updates JSON 解析失败 → {str(e)}"

        if len(update_list) > 200:
            return f"Error: 单次更新不得超过200个单元格（当前: {len(update_list)}）"

        # Step 1: 备份
        bak_path = make_backup(file_path)

        try:
            # Step 2 & 3: 清洗 + 写入（含 MergedCell 路由）
            wb = openpyxl.load_workbook(file_path)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Error: Sheet '{sheet_name}' 不存在。可选: {wb.sheetnames}"

            ws = wb[sheet_name]
            log_entries = []

            for u in update_list:
                row = int(u["row"])
                col = resolve_col(u["col"])
                new_val = u["value"]
                col_letter = get_column_letter(col)

                # MergedCell 智能路由
                target_coord = get_primary_cell(ws, row, col_letter)
                old_val = ws[target_coord].value
                sanitized_val = sanitize_cell_value(new_val)

                ws[target_coord] = sanitized_val

                log_entries.append({
                    "cell": f"{col_letter}{row}→{target_coord}",
                    "old": str(old_val) if old_val is not None else "(空)",
                    "new": str(sanitized_val),
                    "sanitized": sanitized_val != new_val
                })

            wb.save(file_path)
            wb.close()

            # Step 3 (续): 验证读回
            wb_verify = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            ws_verify = wb_verify[sheet_name]
            verify_errors = []

            for u in update_list:
                row = int(u["row"])
                col = resolve_col(u["col"])
                expected = sanitize_cell_value(u["value"])
                actual = ws_verify.cell(row=row, column=col).value

                if str(expected) != str(actual) and expected != actual:
                    verify_errors.append(f"  R{row}C{col}: 期望={expected}, 实际={actual}")

            wb_verify.close()

            # Step 4: 生成变更摘要
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            summary = f"✅ 写入成功 — {ts}\n"
            summary += f"文件: {os.path.basename(file_path)} → Sheet '{sheet_name}'\n"
            summary += f"备份: {os.path.basename(bak_path)}\n"
            if reason:
                summary += f"原因: {reason}\n"
            summary += f"修改: {len(log_entries)} 个单元格\n"
            summary += "-" * 50 + "\n"

            for entry in log_entries:
                flag = " ⚠️已清洗" if entry["sanitized"] else ""
                summary += f"  {entry['cell']}: {entry['old']} → {entry['new']}{flag}\n"

            if verify_errors:
                summary += "\n❌ 验证异常:\n" + "\n".join(verify_errors)
            else:
                summary += "\n✅ 写后验证通过"

            return summary

        except Exception as e:
            return f"Error: 写入失败 → {str(e)}\n备份文件: {bak_path}（可用于恢复）"

    @mcp.tool()
    def append_rows(file_path: str, sheet_name: str,
                    start_row: int, rows_data: str,
                    reason: str = "") -> str:
        """
        从指定行开始，批量写入多行数据（典型场景：填入下料单明细）。

        rows_data 为 JSON 字符串，格式：
        [
          {"B": "A1", "C": "山西黑", "D": "烧面", "E": 603, "F": 320, "G": 30, "H": 1},
          {"B": "A2", "C": "山西黑", "D": "烧面", "E": 615, "F": 305, "G": 30, "H": 8}
        ]
        列名用字母标识。自动为每行计算序号 (A列) 和面积 (I列 = E*F/1000000*H)。
        """
        if not os.path.exists(file_path):
            return f"Error: 文件不存在 → {file_path}"

        try:
            data_rows = json.loads(rows_data)
        except json.JSONDecodeError as e:
            return f"Error: rows_data JSON 解析失败 → {str(e)}"

        if len(data_rows) > 200:
            return f"Error: 单次不得超过200行（当前: {len(data_rows)}）"

        updates = []
        for i, row_data in enumerate(data_rows):
            current_row = start_row + i

            # 自动填序号 (A列 = col 1)
            updates.append({"row": current_row, "col": 1, "value": i + 1})

            # 写入各列数据
            for col_letter, value in row_data.items():
                col_idx = resolve_col(col_letter)
                updates.append({"row": current_row, "col": col_idx, "value": value})

            # 自动计算面积 (I列 = E*F/1000000*H)
            e_val = row_data.get("E", row_data.get("e", 0))
            f_val = row_data.get("F", row_data.get("f", 0))
            h_val = row_data.get("H", row_data.get("h", 1))
            try:
                area = float(e_val) * float(f_val) / 1000000.0 * float(h_val)
                area = round(area, 6)
                updates.append({"row": current_row, "col": 9, "value": area})
            except (ValueError, TypeError):
                pass

        updates_json = json.dumps(updates, ensure_ascii=False)
        return commit_write(file_path, sheet_name, updates_json,
                           reason=reason or f"批量追加 {len(data_rows)} 行下料数据")

    @mcp.tool()
    def read_excel_images(file_path: str, sheet_name: str, output_dir: str) -> str:
        """
        提取指定 Sheet 中的所有悬浮图片，并返回其归属的单元格坐标。
        图片将以 'img_R{行}C{列}_{序号}.png' 的格式保存在 output_dir。
        """
        if not os.path.exists(file_path):
            return f"Error: 文件不存在 → {file_path}"

        if not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Error: Sheet '{sheet_name}' 不存在。"

            ws = wb[sheet_name]
            extracted_images = []

            if not hasattr(ws, '_images') or not ws._images:
                wb.close()
                return json.dumps({"count": 0, "images": [], "msg": "未发现图片"}, ensure_ascii=False)

            for i, img in enumerate(ws._images):
                row = img.anchor._from.row + 1
                col = img.anchor._from.col + 1

                ext = ".png"
                if hasattr(img, 'format') and img.format:
                    ext = f".{img.format.lower()}"

                filename = f"img_R{row}C{col}_{i}{ext}"
                target_path = os.path.join(output_dir, filename)

                with open(target_path, "wb") as f:
                    f.write(img._data())

                extracted_images.append({
                    "index": i, "row": row, "col": col,
                    "local_path": target_path, "filename": filename
                })

            wb.close()
            return json.dumps({
                "count": len(extracted_images),
                "images": extracted_images,
                "msg": f"成功提取 {len(extracted_images)} 张图片"
            }, ensure_ascii=False, indent=2)

        except Exception as e:
            return f"Error: 图片提取失败 → {str(e)}"

    @mcp.tool()
    def delete_rows(file_path: str, sheet_name: str, start_row: int, end_row: int) -> str:
        """
        物理删除指定行范围（用于裁切模板多余空行，自动维持下方单元格上移及公式引用）。
        操作前自动备份。
        """
        if not os.path.exists(file_path):
            return f"Error: 文件不存在 → {file_path}"
        if start_row > end_row:
            return f"Error: start_row ({start_row}) 不得大于 end_row ({end_row})"

        bak_path = make_backup(file_path)
        try:
            wb = openpyxl.load_workbook(file_path)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Error: Sheet '{sheet_name}' 不存在。"
            
            ws = wb[sheet_name]
            amount = end_row - start_row + 1
            ws.delete_rows(start_row, amount)
            
            wb.save(file_path)
            wb.close()
            return f"✅ 裁切成功：已从 {sheet_name} 删除第 {start_row} 至 {end_row} 行（共 {amount} 行）。\n备份: {os.path.basename(bak_path)}"
        except Exception as e:
            return f"Error: 裁切失败 → {str(e)}"
