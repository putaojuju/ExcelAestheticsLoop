"""
services/layout_service.py — ✨ 美学排版通道 (V3 新增)
=====================================================
包含：adjust_column_width, inject_aesthetics_padding, get_layout_metrics

遵循 excel_aesthetics_loop 三原则：
  1. 参数化空间重分配（杜绝平均主义列宽）
  2. 强制美学气囊（AutoFit + 18pt padding）
  3. 原生 COM 物理引擎驱动（禁止 Python 数学猜测行高）
"""

import os
import json
import subprocess
import sys
import tempfile
import openpyxl
from openpyxl.utils import get_column_letter

from core.security import make_backup


def register_layout_tools(mcp):
    """将排版类工具注册到 FastMCP 实例。"""

    @mcp.tool()
    def adjust_column_width(file_path: str, sheet_name: str,
                            col_widths: str, reason: str = "") -> str:
        """
        调整指定列的物理宽度。

        col_widths 为 JSON 字符串，格式：
          {"N": 45.5, "O": 15.0, "B": 28}
        键为列字母，值为 Excel 列宽单位（约等于默认字体下的字符数）。

        操作前自动备份。
        """
        if not os.path.exists(file_path):
            return f"Error: 文件不存在 → {file_path}"

        try:
            width_map = json.loads(col_widths)
        except json.JSONDecodeError as e:
            return f"Error: col_widths JSON 解析失败 → {str(e)}"

        bak_path = make_backup(file_path)

        try:
            wb = openpyxl.load_workbook(file_path)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Error: Sheet '{sheet_name}' 不存在。可选: {wb.sheetnames}"

            ws = wb[sheet_name]
            log = []

            for col_letter, new_width in width_map.items():
                col_letter = col_letter.upper().strip()
                old_width = ws.column_dimensions[col_letter].width
                ws.column_dimensions[col_letter].width = float(new_width)
                log.append(f"  {col_letter}: {old_width} → {new_width}")

            wb.save(file_path)
            wb.close()

            summary = f"✅ 列宽调整成功\n"
            summary += f"文件: {os.path.basename(file_path)} → Sheet '{sheet_name}'\n"
            summary += f"备份: {os.path.basename(bak_path)}\n"
            if reason:
                summary += f"原因: {reason}\n"
            summary += "-" * 40 + "\n"
            summary += "\n".join(log)

            return summary

        except Exception as e:
            return f"Error: 列宽调整失败 → {str(e)}\n备份: {bak_path}"

    @mcp.tool()
    def inject_aesthetics_padding(file_path: str, sheet_name: str,
                                  start_row: int, end_row: int,
                                  padding_pt: float = 18.0,
                                  min_height_pt: float = 35.0,
                                  max_height_pt: float = 250.0) -> str:
        """
        贯彻 excel_aesthetics_loop 三原则的行高美学引擎（V3.2 隔离 + Stream IPC 版）。

        执行流程：通过 Stdin IPC 信号唤起 core/com_bridge.py 执行 Native COM 操作。
        """
        if not os.path.exists(file_path):
            return f"Error: 文件不存在 → {file_path}"

        bak_path = make_backup(file_path)
        bridge_script = os.path.join(os.path.dirname(os.path.dirname(__file__)), "core", "com_bridge.py")
        
        payload = {
            "cmd": "padding",
            "file_path": file_path,
            "sheet_name": sheet_name,
            "start_row": start_row,
            "end_row": end_row,
            "padding_pt": padding_pt,
            "min_height": min_height_pt,
            "max_height": max_height_pt
        }

        try:
            result = subprocess.run(
                [sys.executable, bridge_script],
                input=json.dumps(payload),
                capture_output=True,
                text=True,
                encoding='utf-8'
            )
            
            if result.returncode != 0:
                return f"Error: 桥接调度失败 → {result.stderr}"
            
            data = json.loads(result.stdout)
            if not data.get("success"):
                return f"Error: 物理引擎执行失败 → {data.get('error')}"

            summary = f"✅ 美学气囊注入完毕 (V3.2 流通信模式)\n"
            summary += f"文件: {os.path.basename(file_path)}\n"
            summary += f"范围: Row {start_row} ~ {end_row}\n"
            if data.get("clamped"):
                summary += f"⚠️ 列宽建议：以下行触及了 {max_height_pt}pt 上限，可能存在严重截断: {data['clamped']}"
            
            return summary

        except Exception as e:
            return f"Error: 调度异常 → {str(e)}"

    @mcp.tool()
    def get_layout_metrics(file_path: str, sheet_name: str) -> str:
        """[The Inspector] 扫描合并单元格和列宽度，用于解决排版中的「空间盲区」问题。"""
        if not os.path.exists(file_path):
            return f"Error: 文件不存在 → {file_path}"
        try:
            # merged_cells 不支持 read_only 模式
            wb = openpyxl.load_workbook(file_path, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Error: Sheet '{sheet_name}' 不存在。"
            ws = wb[sheet_name]
            merged = [str(r) for r in ws.merged_cells.ranges]
            metrics = {
                "merged_cells": merged,
                "column_widths": {
                    get_column_letter(i): ws.column_dimensions[get_column_letter(i)].width
                    for i in range(1, ws.max_column + 1)
                }
            }
            wb.close()
            return json.dumps(metrics, ensure_ascii=False, indent=2)
        except Exception as e:
            return f"Error: {str(e)}"
