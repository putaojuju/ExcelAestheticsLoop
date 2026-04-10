"""
services/vision_service.py — 视觉审计通道
==========================================
包含：smart_audit_read, get_visual_slice

依赖 core/native_rip.py 的物理渲染引擎。
"""

import os
import json
from datetime import datetime
import openpyxl

from config import CACHE_DIR, AUDIT_KEYWORDS
from core.excel_mapping import int_to_col_letter
from core.native_rip import render_range_to_png


def register_vision_tools(mcp):
    """将视觉审计类工具注册到 FastMCP 实例。"""

    @mcp.tool()
    def smart_audit_read(file_path: str, sheet_name: str) -> str:
        """
        读取指定 Sheet 的 Schema，并自动捕获「高危列 (备注/单号/附件)」的高清视觉快照。
        返回 JSON 包含数据结构及生成的图片路径。
        """
        if not os.path.exists(file_path):
            return f"Error: 文件不存在 → {file_path}"

        if not os.path.exists(CACHE_DIR):
            os.makedirs(CACHE_DIR, exist_ok=True)

        ts = datetime.now().strftime("%H%M%S")
        report = {
            "file": os.path.basename(file_path),
            "sheet": sheet_name,
            "timestamp": ts,
            "columns": [],
            "visual_cache": []
        }

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return f"Error: Sheet '{sheet_name}' 不存在。"

            ws = wb[sheet_name]

            # 增强型表头扫描：扫描前 5 行 × 最大 30 列
            column_meta = {}
            scan_cols = max(30, ws.max_column if ws.max_column else 0)
            for r in range(1, 6):
                for c in range(1, scan_cols + 1):
                    val = ws.cell(row=r, column=c).value
                    if val:
                        val_str = str(val).lower()
                        if c not in column_meta:
                            column_meta[c] = {"header": str(val), "needs_visual": False}
                        if any(k in val_str for k in AUDIT_KEYWORDS):
                            column_meta[c]["needs_visual"] = True
                            column_meta[c]["header"] = str(val)

            for c in range(1, (ws.max_column or 1) + 1):
                col_letter = int_to_col_letter(c)
                meta = column_meta.get(c, {"header": f"Column_{c}", "needs_visual": False})
                report["columns"].append({
                    "letter": col_letter,
                    "header": meta["header"],
                    "needs_visual": meta["needs_visual"]
                })
            wb.close()

            # 视觉层 RIP（针对高危列）
            for col in report["columns"]:
                if col["needs_visual"]:
                    letter = col["letter"]
                    safe_h = "".join(c for c in col["header"] if c.isalnum())[:10]
                    out_name = f"audit_{ts}_{letter}_{safe_h}.png"
                    out_path = os.path.join(CACHE_DIR, out_name)

                    if render_range_to_png(file_path, sheet_name,
                                           f"{letter}1:{letter}30", out_path):
                        col["visual_cache_path"] = out_path
                        report["visual_cache"].append(out_path)

            return json.dumps(report, ensure_ascii=False, indent=2)
        except Exception as e:
            return f"Error: 审计读取失败 (V3.1) → {str(e)}"

    @mcp.tool()
    def get_visual_slice(file_path: str, sheet_name: str, range_str: str) -> str:
        """
        对 Excel 指定区域进行高清截图渲染。
        range_str 可以是 'B1:C20' 或单列字母 'O'。
        """
        if not os.path.exists(CACHE_DIR):
            os.makedirs(CACHE_DIR, exist_ok=True)

        # 智能补全：单列字母自动展开为 40 行
        if ":" not in range_str and len(range_str) <= 3:
            range_str = f"{range_str}1:{range_str}40"

        ts = datetime.now().strftime("%H%M%S")
        out_path = os.path.join(CACHE_DIR, f"rip_{ts}_{range_str.replace(':', '_')}.png")

        if render_range_to_png(file_path, sheet_name, range_str, out_path):
            return f"OK: 视觉切片已生成 → {out_path}"
        else:
            return "Error: 渲染失败，请检查 Excel 状态或区域有效性。"
